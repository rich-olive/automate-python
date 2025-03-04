import openpyxl
from openpyxl.styles import Font, PatternFill

# 1. Load the Spreadsheet
employee_sales = openpyxl.load_workbook("Employee Sales.xlsx")

# 2. Take User Input to Specify Worksheets and Ranges
# use static data for now
dataset1 = employee_sales["Dataset 1"]
dataset2 = employee_sales["Dataset 2"]

#b2:e21

# compare names
# a2:a21
ds1 = dataset1["A2:A21"]
ds2 = dataset2["A2:A21"]


no_match_style = PatternFill(start_color="0000FF", end_color="0000FF", patternType="solid")

# 4. Create New Worksheet and Write Differences:
del employee_sales["diff"]
employee_sales.create_sheet("diff")

errors = []
headers = ["Row", "Column", "Value 1", "Value 2", "Delta"]
errors.append(headers)

for row in dataset1.iter_rows():

    for cell in row:
        current_cell = cell.coordinate
        value1 = dataset1[current_cell].value
        value2 = dataset2[current_cell].value
        # percentage_change = ((value1 - value2) / value2) * 100
        if value1 != value2:
            if type(value1) == int and type(value2) == int:
                errors.append([cell.row, dataset1[cell.column_letter + "1"].value, value1, value2, ((value1 - value2) / value1) * 100])
            else:
                errors.append([cell.row, dataset1[cell.column_letter + "1"].value, value1, value2, ""])
# 3. Perform Comparison Between Specified Cell Ranges
errors_range = "A1:E" + str(len(errors))
rng = employee_sales["diff"][errors_range]

for i in range(len(errors)):
    row = errors[i]
    for j in range(len(row)):
        val = row[j]
        rng[i][j].value = val

# perform comparison calculation for numerical data differences
#

employee_sales.save("Employee Sales.xlsx")