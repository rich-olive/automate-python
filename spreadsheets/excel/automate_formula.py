import openpyxl

wb = openpyxl.load_workbook("Formulas, Fonts, and Fills.xlsx")

ws = wb["Ratings"]

formula = "=AVERAGE(B2:E2)"

rng = ws["B2:E11"]

ws["F1"].value = "Average Rating"
# ws["F2"] = formula

# basically building out the formula cell references 'dynamically'
for index, row in enumerate(rng):
    current_row = str(index + 2)

    formula_cell = "F" + current_row

    formula = f"=AVERAGE(B{current_row}:E{current_row})"

    ws[formula_cell] = formula

    for cell in row:
        val = cell.value
        if type(val) is not int:
            cell.value = ""
        elif val < 0:
            cell.value = 0
        elif val > 10:
            cell.value = 10

wb.save("Formulas, Fonts, and Fills.xlsx")

# it's easier to write formulae in python
# e.g. write a program to output the average of a number of values
# but if our spreadsheets are being shared with others
# or potentially audited, the calculations need to be transparent
# hence it's better to write formulae in an excel way