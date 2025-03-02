import openpyxl

wb = openpyxl.load_workbook("my_workbook.xlsx")

ws = wb["Sheet"]

c = ws["A1"]

c_value = c.value

print(c_value) # Hello Excel

# Example of it all chained together
ws2 = wb["Sheet2"]["A1"].value
print(ws2) # Hi Olivia

# Reverse a value:
reverse_text = c_value[::-1]
ws["A2"] = reverse_text

# WRITING TO A CELL WITHOUT USING CELL LETTER
# it can become very difficult to target cells based on their letter value
ws.cell(row=11, column=11, value="Greetings")
# retrieving a value
cell_val2 = ws.cell(row=11, column=11).value

print(cell_val2)

wb.save("my_workbook.xlsx")