import openpyxl
from openpyxl.reader.excel import load_workbook

book_data = [
    ["Conclave","Robert Harris","Fiction"],
    ["Goodnight, Mister Tom", "Michelle Magorian", "Fiction"],
    ["Davita's Harp", "Chaim Potok", "Fiction"]
]

wb = load_workbook("books.xlsx")
ws = wb.active
# str(len(book_data))
cell_range = "A5:C7"
rng = ws[cell_range]
print(len(book_data))

for i in range(len(book_data)):
    row = book_data[i]
    for j in range(len(row)):
        val = row[j]
        rng[i][j].value = val

# we use the range (ranges.py) in the for loop
# but what did we use it for in writing_multiple.py ??

# for i in range(len(book_data)):
#     row = book_data[i]
#     for j in range(len(row)):
#         val = row[j]
#
#         cell_range[i][j].value = val
#         # wb["books"]["A5:C7"][1][2].value = book_data[1][2]

# print(ws[new_cell_range])

# ws2 = wb["Sheet2"]["A1"].value
# c = ws["C7"].value
#
# c.value = "another book"

# ws["C7"] = "another book"

wb.save("books.xlsx")

# the error (out of range etc) was because the way the tutorial taught me to dynamically add a range
# was to write a start position, then base the end position on the length of the data (converted to a string)
# but in this instance, it was essentially writing A6:C3 because there was only 3 items in the new data list
