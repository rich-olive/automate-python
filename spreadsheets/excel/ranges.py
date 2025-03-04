import openpyxl

wb = openpyxl.load_workbook("Employee Ratings.xlsx")

ws = wb["Ratings"]

# top left to top right
# B2 - E11
# for one dimensional: B2-B11

# rng2 = ws["B2:B6"]
# print(rng2)
rng = ws["B2:E11"]
print(rng) #((<Cell 'Ratings'.B2>, <Cell 'Ratings'.C2>, <Cell 'Ratings'.D2>, <Cell 'Ratings'.E2>),...
# a tuple containing more tuples
# B2, C2, D2, E2
# similar for 1D data, it would just return B2 in the first nested tuple
# B3 in the second nested tuple
# B4 in the third, etc

# so we need to use nested for loops to handle the data
# the outer tuple is the row
# the inner tuple is the cell
for row in rng:
    for cell in row:
        val = cell.value
        # print(cell.value)
        # 3
        # 100
        # 9 etc
        if type(val) is not int:
            cell.value = ""
        elif val < 0:
            cell.value = 0
        elif val > 10:
            cell.value = 10

wb.save("Employee Ratings.xlsx")